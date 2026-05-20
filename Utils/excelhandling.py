def test_excelhandling():
    import openpyxl
    workbook = openpyxl.load_workbook("testData/creds.xlsx")
    sheet = workbook.active
    values=[]
   # print(sheet.cell(row=1, column=1).value)
    for i in range(1, sheet.max_row + 1):
        row_values = []
        for j in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=i, column=j).value
            row_values.append(cell_value)
        values.append(row_values)